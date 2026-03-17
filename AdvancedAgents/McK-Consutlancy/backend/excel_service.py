import os
import re
from html import escape

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

TABLE_SEP_RE = re.compile(r"^\|[\s\-\|:]+\|$")
TABLE_ROW_RE = re.compile(r"^\|(.+)\|$")
HEADING_RE = re.compile(r"^#{1,6}\s+(.+)$")


def parse_markdown_tables(text: str) -> list:
    """Parse all GFM pipe tables from markdown text.

    Returns a list of dicts: {"title": str, "headers": [str], "rows": [[str]]}
    """
    if not text:
        return []

    tables = []
    lines = text.splitlines()
    current_heading = ""
    i = 0

    while i < len(lines):
        line = lines[i].rstrip()

        heading_match = HEADING_RE.match(line)
        if heading_match:
            current_heading = heading_match.group(1).strip()
            i += 1
            continue

        # Detect GFM table: pipe row followed by separator row
        if TABLE_ROW_RE.match(line) and i + 1 < len(lines):
            next_line = lines[i + 1].rstrip()
            if TABLE_SEP_RE.match(next_line):
                # Parse header row
                headers = [c.strip() for c in line.strip("|").split("|")]
                i += 2  # skip header + separator

                # Collect data rows
                rows = []
                while i < len(lines):
                    data_line = lines[i].rstrip()
                    if not TABLE_ROW_RE.match(data_line):
                        break
                    cells = [c.strip() for c in data_line.strip("|").split("|")]
                    # Pad to header width
                    while len(cells) < len(headers):
                        cells.append("")
                    rows.append(cells[:len(headers)])
                    i += 1

                tables.append({
                    "title": current_heading,
                    "headers": headers,
                    "rows": rows,
                })
                continue

        i += 1

    return tables


def _safe_sheet_name(title: str, existing: list) -> str:
    """Return a safe, unique Excel sheet name (max 31 chars, no special chars)."""
    # Strip illegal chars
    safe = re.sub(r'[\\/?*\[\]:]', '', title).strip()[:28] or "Sheet"
    # Deduplicate
    candidate = safe
    n = 1
    while candidate in existing:
        candidate = f"{safe}_{n}"
        n += 1
    return candidate


def save_tables_to_excel(tables: list, analysis_id: str) -> str:
    """Save parsed tables to an xlsx file. Returns the file path."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="0A2342")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    even_fill = PatternFill("solid", fgColor="EBF0FA")

    sheet_names = []

    if not tables:
        ws = wb.create_sheet("No Tables Found")
        ws["A1"] = "No tabular data was detected in this analysis."
        sheet_names.append("No Tables Found")
    else:
        for table in tables:
            sheet_name = _safe_sheet_name(table["title"] or "Table", sheet_names)
            sheet_names.append(sheet_name)
            ws = wb.create_sheet(sheet_name)

            headers = table["headers"]
            rows = table["rows"]

            # Write header row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            # Write data rows
            for row_idx, row in enumerate(rows, start=2):
                fill = even_fill if row_idx % 2 == 0 else None
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(size=9)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True)
                    if fill:
                        cell.fill = fill

            # Auto column widths
            for col_idx, header in enumerate(headers, start=1):
                col_values = [header] + [r[col_idx - 1] for r in rows if col_idx - 1 < len(r)]
                max_len = max((len(str(v)) for v in col_values), default=8)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    path = os.path.join(EXPORTS_DIR, f"analysis_{analysis_id}_tables.xlsx")
    wb.save(path)
    return path


def render_excel_tables_as_html(excel_path: str) -> str:
    """Read an xlsx file and return inline-styled HTML tables."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    parts = []

    for sheet in wb.worksheets:
        parts.append(f"<p><strong>{escape(sheet.title)}</strong></p>")
        parts.append(
            '<table style="width:100%;border-collapse:collapse;margin:10px 0;font-size:9pt;">'
        )

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                row_style = 'background:#0A2342;color:#fff;font-weight:bold;'
            else:
                row_style = 'background:#EBF0FA;' if row_idx % 2 == 1 else 'background:#FFFFFF;'

            parts.append("<tr>")
            for cell_val in row:
                value = escape(str(cell_val)) if cell_val is not None else ""
                parts.append(
                    f'<td style="{row_style}border:1px solid #ddd;padding:4px 6px;">{value}</td>'
                )
            parts.append("</tr>")

        parts.append("</table>")

    wb.close()
    return "\n".join(parts)
